# -*- coding: utf-8 -*-
"""
Rolling Range Volatility - 滚动极差波动率
=========================================
复用 backtest_main.py 的数据加载与逐bar上下文风格，计算多个分钟周期的：
    (窗口最高价 - 窗口最低价) / 窗口开盘价

输出：
1) 按时间顺序的折线图
2) 按振幅值统计的直方图
"""

import os
import time
import numpy as np
import pandas as pd
from matplotlib import pyplot as plt
from openpyxl import load_workbook

from backtest_main import load_data, BarContext


def compute_rrv_by_bar(quote: pd.DataFrame, window_bars: int,
                       period_label: str) -> pd.DataFrame:
    """
    逐bar计算滚动极差波动率。
    RRV = (max(high) - min(low)) / open_of_window_start
    """
    signal = pd.DataFrame(
        index=quote.index,
        columns=['date', 'rrv_up', 'rrv_down'],
        dtype=float
    )
    n = len(quote)
    print(f'[RRV] {period_label}: start, window={window_bars} bars, total={n}')

    progress_step = max(1, n // 10)

    for ii, index in enumerate(quote.index):
        ctx = BarContext(
            quote=quote,
            signal=signal,
            index=index,
            integer_index=ii,
        )

        signal.at[index, 'date'] = ctx.quote.iat[ii, 0]

        if ii + 1 >= window_bars:
            start = ii - window_bars + 1
            win = ctx.quote.iloc[start:ii + 1]
            win_open = win['open'].iloc[0]
            win_high = win['high'].max()
            win_low = win['low'].min()
            rrv_up = (win_high - win_open) / win_open if win_open != 0 else np.nan
            rrv_down = (win_open - win_low) / win_open if win_open != 0 else np.nan
            signal.at[index, 'rrv_up'] = rrv_up
            signal.at[index, 'rrv_down'] = rrv_down

        if ((ii + 1) % progress_step == 0) or (ii + 1 == n):
            pct = round((ii + 1) / n * 100, 1)
            print(f'[RRV] {period_label}: {ii + 1}/{n} ({pct}%)')

    return signal.dropna(subset=['rrv_up', 'rrv_down']).copy()


def plot_rrv_line(rrv_df: pd.DataFrame, period_label: str, metric_col: str,
                  direction_tag: str, out_dir: str):
    """按时间顺序画 RRV 折线图。"""
    fig = plt.figure(figsize=(18, 7))
    ax = fig.add_subplot(111)

    dates = pd.to_datetime(rrv_df['date'], errors='coerce')
    y = rrv_df[metric_col].to_numpy(dtype=float)
    mask = dates.notna()
    dates = dates[mask]
    y = y[mask.to_numpy()]

    # 用等间距 bar 序号作 x 轴，避免日期轴在缺失时段产生空白
    x = np.arange(len(y))
    ax.plot(x, y, linewidth=0.8, color='tab:blue')
    ax.set_title(f'Rolling Range Volatility ({direction_tag}) - {period_label} (Time Order)')
    ax.set_xlabel('Date (MM-DD)')
    ax.set_ylabel(metric_col)
    ax.grid(alpha=0.2)
    tick_count = min(14, max(6, len(x) // 50000 + 6))
    tick_idx = np.linspace(0, len(x) - 1, num=tick_count, dtype=int)
    tick_idx = np.unique(tick_idx)
    tick_labels = dates.iloc[tick_idx].dt.strftime('%m-%d')
    ax.set_xticks(tick_idx)
    ax.set_xticklabels(tick_labels, rotation=30, ha='right')

    save_path = os.path.join(out_dir, f'rrv_{period_label}_{direction_tag}_line.png')
    fig.savefig(save_path, dpi=220, bbox_inches='tight')
    plt.close(fig)


def plot_rrv_hist(rrv_df: pd.DataFrame, period_label: str, metric_col: str,
                  direction_tag: str, out_dir: str):
    """按振幅大小排序后画直方图。"""
    fig = plt.figure(figsize=(12, 7))
    ax = fig.add_subplot(111)

    vals = rrv_df[metric_col].to_numpy(dtype=float)
    vals = vals[np.isfinite(vals)]
    if len(vals) == 0:
        return

    # 自适应横轴：避免少量极端值把右侧拉出大面积空白
    p997 = float(np.quantile(vals, 0.997))
    vmax = float(vals.max())
    upper = p997 if p997 > 0 else vmax
    upper = max(upper * 1.05, 1e-6)  # 留少量边距

    vals_in = vals[vals <= upper]
    vals_out = vals[vals > upper]
    bins = 120

    ax.hist(vals_in, bins=bins, color='tab:blue', alpha=0.85, edgecolor='black')
    ax.set_xlim(0, upper)
    ax.set_title(f'Rolling Range Volatility ({direction_tag}) - {period_label} (Histogram)')
    ax.set_xlabel(metric_col)
    ax.set_ylabel('Count')
    ax.grid(alpha=0.2)

    if len(vals_out) > 0:
        ax.text(
            0.99, 0.97,
            f'clipped > {upper:.6f}: {len(vals_out)}',
            transform=ax.transAxes,
            ha='right', va='top', fontsize=9,
            bbox=dict(boxstyle='round', fc='white', alpha=0.75)
        )
    print(f'[RRV] {period_label} hist range: 0 ~ {upper:.6f}, clipped={len(vals_out)}')

    save_path = os.path.join(out_dir, f'rrv_{period_label}_{direction_tag}_hist.png')
    fig.savefig(save_path, dpi=220, bbox_inches='tight')
    plt.close(fig)


def build_rrv_stats(rrv_df: pd.DataFrame, period_label: str, metric_col: str,
                    direction_tag: str, window_bars: int,
                    bar_seconds: int) -> pd.DataFrame:
    """构建 RRV 统计摘要（用于 Excel 输出）。"""
    vals = pd.to_numeric(rrv_df[metric_col], errors='coerce')
    vals = vals[np.isfinite(vals)]
    dates = pd.to_datetime(rrv_df['date'], errors='coerce')

    if len(vals) == 0:
        stats = [
            ('period', period_label),
            ('direction', direction_tag),
            ('metric_col', metric_col),
            ('window_bars', window_bars),
            ('bar_seconds', bar_seconds),
            ('sample_count', 0),
        ]
        return pd.DataFrame(stats, columns=['metric', 'value'])

    p = vals.quantile([0.01, 0.05, 0.25, 0.50, 0.75, 0.95, 0.99])
    stats = [
        ('period', period_label),
        ('direction', direction_tag),
        ('metric_col', metric_col),
        ('window_bars', int(window_bars)),
        ('bar_seconds', int(bar_seconds)),
        ('sample_count', int(len(vals))),
        ('start_date', str(dates.min())),
        ('end_date', str(dates.max())),
        ('mean', float(vals.mean())),
        ('std', float(vals.std(ddof=1)) if len(vals) > 1 else 0.0),
        ('min', float(vals.min())),
        ('p01', float(p.loc[0.01])),
        ('p05', float(p.loc[0.05])),
        ('p25', float(p.loc[0.25])),
        ('p50', float(p.loc[0.50])),
        ('p75', float(p.loc[0.75])),
        ('p95', float(p.loc[0.95])),
        ('p99', float(p.loc[0.99])),
        ('max', float(vals.max())),
        ('gt_0.01_count', int((vals > 0.01).sum())),
        ('gt_0.01_ratio', float((vals > 0.01).mean())),
    ]
    return pd.DataFrame(stats, columns=['metric', 'value'])


if __name__ == '__main__':
    start_time = time.time()

    # 数据源（与现有策略脚本保持一致）
    folder_path = r"F:\Data\XAGUSD\\"
    file_name = "xagusd_30s_all"

    # 可选: 'up' / 'down' / 'both'
    run_direction = 'up'

    quote, _, bar_seconds = load_data(folder_path, file_name)

    # 可按需缩小区间，例如：
    # quote = quote[(quote.index > 30000) & (quote.index < 35000)].reset_index(drop=True)

    periods_min = [1, 3, 5, 15, 30, 60]
    periods_min = [90, 120, 150, 180]
    # periods_min = [240, 300, 480, 1440]
    out_dir = f'./{file_name} rolling range volatility'
    os.makedirs(out_dir, exist_ok=True)

    print(f'[RRV] output dir: {out_dir}')
    print(f'[RRV] periods(min): {periods_min}')
    print(f'[RRV] direction mode: {run_direction}')
    excel_data_by_period = {}

    if run_direction not in {'up', 'down', 'both'}:
        raise ValueError("run_direction must be 'up' or 'down' or 'both'")

    if run_direction == 'up':
        selected_metrics = [('rrv_up', 'up')]
    elif run_direction == 'down':
        selected_metrics = [('rrv_down', 'down')]
    else:
        selected_metrics = [('rrv_up', 'up'), ('rrv_down', 'down')]

    for m in periods_min:
        window_bars = int((m * 60) / bar_seconds)
        period_label = f'{m}min'

        rrv_df = compute_rrv_by_bar(quote, window_bars, period_label)
        if len(rrv_df) == 0:
            print(f'[RRV] {period_label}: no valid values, skip plots.')
            continue

        for metric_col, direction_tag in selected_metrics:
            stats_df = build_rrv_stats(
                rrv_df, period_label, metric_col, direction_tag,
                window_bars, bar_seconds
            )
            excel_data_by_period[f'{period_label}_{direction_tag}'] = stats_df

            plot_rrv_line(rrv_df, period_label, metric_col, direction_tag, out_dir)
            plot_rrv_hist(rrv_df, period_label, metric_col, direction_tag, out_dir)

        print(f'[RRV] {period_label}: done, values={len(rrv_df)}')

    if len(excel_data_by_period) > 0:
        excel_path = os.path.join(out_dir, f'{file_name}_rrv_data_{run_direction}.xlsx')
        print(f'[RRV] writing excel: {excel_path}')
        if os.path.exists(excel_path):
            wb = load_workbook(excel_path)
            existing_sheets = set(wb.sheetnames)
            wb.close()
            mode = 'a'
        else:
            existing_sheets = set()
            mode = 'w'

        with pd.ExcelWriter(excel_path, engine='openpyxl', mode=mode) as writer:
            for period_label, export_df in excel_data_by_period.items():
                sheet_name = period_label
                suffix = 1
                while sheet_name in existing_sheets:
                    sheet_name = f'{period_label}_{suffix}'
                    suffix += 1

                export_df.to_excel(writer, sheet_name=sheet_name, index=False)
                existing_sheets.add(sheet_name)
                print(f'[RRV] excel sheet done: {sheet_name}, rows={len(export_df)}')

    print(f'[RRV] total time: {round(time.time() - start_time, 2)}s')
